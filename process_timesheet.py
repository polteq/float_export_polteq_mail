#!/usr/bin/env python3
"""
Timesheet Automation Script

Processes CSV files from Float/People and converts them to Excel and PDF formats.
Automatically organizes files into appropriate folders.

Usage:
    python process_timesheet.py
"""

import os
import sys
import re
import logging
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import subprocess
import shutil

# Configuration
SCRIPT_DIR = Path(__file__).parent
IMPORT_FOLDER = SCRIPT_DIR / "Import Urenstaat"
CONVERTED_FOLDER = SCRIPT_DIR / "converted"
ONEDRIVE_FOLDER = Path("/Users/pl-tq-261/Library/CloudStorage/OneDrive-Polteq/Urenstaat")
LOG_FOLDER = SCRIPT_DIR / "logs"
CONFIG_FILE = SCRIPT_DIR / "email_config.json"

# Email configuration (can be overridden by email_config.json)
DEFAULT_EMAIL_CONFIG = {
    "enabled": True,
    "recipient": "uren@polteq.com",
    "cc": "",
    "subject": "Urenstaat Jorre - Signify {month_year}",
    "body": "Zie bijlage voor de urenstaat van {month_year}.\n\n"
}

# Ensure folders exist
IMPORT_FOLDER.mkdir(exist_ok=True)
CONVERTED_FOLDER.mkdir(exist_ok=True)
ONEDRIVE_FOLDER.mkdir(parents=True, exist_ok=True)
LOG_FOLDER.mkdir(exist_ok=True)

# Setup logging
log_file = LOG_FOLDER / f"timesheet_processing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


def extract_date_from_filename(filename):
    """
    Extract end date from filename pattern: People-LoggedTime-YYYYMMDD-YYYYMMDD.csv
    Returns the end date (second date) as a datetime object.
    """
    pattern = r'(.+?)-LoggedTime-(\d{8})-(\d{8}).*?\.csv'
    match = re.search(pattern, filename)
    
    if not match:
        raise ValueError(f"Filename '{filename}' does not match expected pattern")
    
    person_name = match.group(1)
    start_date_str = match.group(2)
    end_date_str = match.group(3)
    
    logger.info(f"Extracted dates from filename: person={person_name}, start={start_date_str}, end={end_date_str}")
    
    # Parse end date
    end_date = datetime.strptime(end_date_str, '%Y%m%d')
    return end_date


def read_csv_with_multiple_delimiters(file_path):
    """
    Read CSV file, trying both comma and tab delimiters.
    Returns a pandas DataFrame.
    """
    logger.info(f"Reading CSV file: {file_path}")
    
    # Try comma delimiter first
    try:
        df = pd.read_csv(file_path, delimiter=',')
        if len(df.columns) > 1:
            logger.info(f"Successfully read CSV with comma delimiter ({len(df.columns)} columns)")
            return df
    except Exception as e:
        logger.warning(f"Failed to read with comma delimiter: {e}")
    
    # Try tab delimiter
    try:
        df = pd.read_csv(file_path, delimiter='\t')
        if len(df.columns) > 1:
            logger.info(f"Successfully read CSV with tab delimiter ({len(df.columns)} columns)")
            return df
    except Exception as e:
        logger.warning(f"Failed to read with tab delimiter: {e}")
    
    # Try auto-detection
    try:
        df = pd.read_csv(file_path, sep=None, engine='python')
        logger.info(f"Successfully read CSV with auto-detected delimiter ({len(df.columns)} columns)")
        return df
    except Exception as e:
        logger.error(f"Failed to read CSV with any delimiter: {e}")
        raise


def convert_to_excel(df, output_path):
    """
    Convert DataFrame to Excel file with basic formatting.
    """
    logger.info(f"Converting to Excel: {output_path}")
    
    # Write to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Timesheet')
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Timesheet']
        
        # Apply basic formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Format header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    logger.info(f"Excel file created successfully: {output_path}")


def create_outlook_email(excel_path, pdf_path, month_year):
    """
    Create a new email in Outlook with the Excel and PDF files attached.
    Uses AppleScript to automate Outlook.
    """
    logger.info("Creating Outlook email with attachments...")
    
    # Load email configuration
    email_config = DEFAULT_EMAIL_CONFIG.copy()
    if CONFIG_FILE.exists():
        try:
            import json
            with open(CONFIG_FILE, 'r') as f:
                user_config = json.load(f)
                email_config.update(user_config)
            logger.info(f"Loaded email configuration from {CONFIG_FILE}")
        except Exception as e:
            logger.warning(f"Could not load email config: {e}. Using defaults.")
    
    if not email_config.get("enabled", True):
        logger.info("Email automation is disabled in config")
        return False
    
    # Format subject and body with month/year
    subject = email_config["subject"].format(month_year=month_year)
    body = email_config["body"].format(month_year=month_year)
    recipient = email_config.get("recipient", "")
    cc = email_config.get("cc", "")
    
    try:
        # Build AppleScript to create email with attachments
        applescript = f'''
        tell application "Microsoft Outlook"
            set newMessage to make new outgoing message with properties {{subject:"{subject}", content:"{body}"}}
            
            -- Add recipient if specified
            {f'make new recipient at newMessage with properties {{email address:{{address:"{recipient}"}}}}' if recipient else ''}
            
            -- Add CC if specified
            {f'make new cc recipient at newMessage with properties {{email address:{{address:"{cc}"}}}}' if cc else ''}
            
            -- Attach files
            make new attachment at newMessage with properties {{file:POSIX file "{excel_path}"}}
            make new attachment at newMessage with properties {{file:POSIX file "{pdf_path}"}}
            
            -- Open the email (don't send automatically)
            open newMessage
            activate
        end tell
        '''
        
        result = subprocess.run(
            ['osascript', '-e', applescript],
            capture_output=True,
            text=True,
            timeout=10
        )
        
        if result.returncode == 0:
            logger.info("✓ Email created successfully in Outlook")
            return True
        else:
            logger.error(f"Failed to create Outlook email: {result.stderr}")
            return False
            
    except Exception as e:
        logger.error(f"Error creating Outlook email: {e}")
        return False
def convert_csv_to_pdf(csv_path, pdf_path):
    """
    Convert CSV file to PDF using Numbers.app (macOS native).
    This produces the same output as manually exporting from Numbers.
    """
    logger.info(f"Converting CSV to PDF: {pdf_path}")
    
    # First, try using Numbers.app via AppleScript (best quality, matches manual export)
    try:
        applescript = f'''
        tell application "Numbers"
            set theDoc to open POSIX file "{csv_path}"
            delay 1
            export theDoc to POSIX file "{pdf_path}" as PDF
            close theDoc without saving
        end tell
        '''
        
        result = subprocess.run(
            ['osascript', '-e', applescript],
            capture_output=True,
            text=True,
            timeout=30
        )
        
        if result.returncode == 0 and pdf_path.exists():
            logger.info(f"PDF created successfully using Numbers: {pdf_path}")
            return True
        else:
            logger.warning(f"Numbers export failed: {result.stderr}")
    except Exception as e:
        logger.warning(f"Error using Numbers for PDF conversion: {e}")
    
    # Fallback: Try LibreOffice
    logger.info("Trying LibreOffice as fallback...")
    libreoffice_paths = [
        '/Applications/LibreOffice.app/Contents/MacOS/soffice',
        '/usr/local/bin/soffice',
        '/opt/homebrew/bin/soffice'
    ]
    
    soffice_path = None
    for path in libreoffice_paths:
        if os.path.exists(path):
            soffice_path = path
            break
    
    if soffice_path:
        try:
            cmd = [
                soffice_path,
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', str(pdf_path.parent),
                str(csv_path)
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
            
            if result.returncode == 0:
                temp_pdf = pdf_path.parent / f"{csv_path.stem}.pdf"
                
                if temp_pdf.exists():
                    if temp_pdf != pdf_path:
                        if pdf_path.exists():
                            os.remove(pdf_path)
                        os.rename(temp_pdf, pdf_path)
                    logger.info(f"PDF created successfully using LibreOffice: {pdf_path}")
                    return True
        except Exception as e:
            logger.warning(f"LibreOffice conversion failed: {e}")
    
    # Final fallback: Create a simple text-based PDF using reportlab
    logger.info("Using Reportlab as final fallback...")
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        
        # Read the CSV file
        df = pd.read_csv(csv_path)
        
        # Create PDF in landscape
        doc = SimpleDocTemplate(str(pdf_path), pagesize=landscape(A4))
        elements = []
        
        # Add title
        styles = getSampleStyleSheet()
        title = Paragraph(f"<b>Urenstaat - {pdf_path.stem}</b>", styles['Title'])
        elements.append(title)
        
        # Convert DataFrame to list for table
        data = [df.columns.tolist()] + df.values.tolist()
        
        # Create table with smaller font to fit
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        logger.info(f"Simple PDF created successfully: {pdf_path}")
        return True
        
    except ImportError:
        logger.error("reportlab not installed. Cannot create PDF.")
        return False
    except Exception as e:
        logger.error(f"Error creating simple PDF: {e}")
        return False


def process_timesheet_file(csv_file):
    """
    Process a single timesheet CSV file.
    """
    logger.info(f"\n{'='*60}")
    logger.info(f"Processing file: {csv_file.name}")
    logger.info(f"{'='*60}")
    
    try:
        # Extract date from filename
        end_date = extract_date_from_filename(csv_file.name)
        month_year = end_date.strftime('%m-%Y')
        
        # Read CSV
        df = read_csv_with_multiple_delimiters(csv_file)
        logger.info(f"Loaded {len(df)} rows and {len(df.columns)} columns")
        
        # Generate output filenames
        base_filename = f"Urenstaat Signify-Jorre van Munster {month_year}"
        excel_filename = f"{base_filename}.xlsx"
        pdf_filename = f"{base_filename}.pdf"
        
        # Create temporary files in script directory first
        temp_excel_path = SCRIPT_DIR / excel_filename
        temp_pdf_path = SCRIPT_DIR / pdf_filename
        
        # Convert to Excel
        convert_to_excel(df, temp_excel_path)
        
        # Convert CSV directly to PDF (better formatting than from Excel)
        pdf_success = convert_csv_to_pdf(csv_file, temp_pdf_path)
        
        # Move files to OneDrive
        final_excel_path = ONEDRIVE_FOLDER / excel_filename
        final_pdf_path = ONEDRIVE_FOLDER / pdf_filename
        
        excel_ready_path = temp_excel_path
        pdf_ready_path = temp_pdf_path
        
        try:
            shutil.move(str(temp_excel_path), str(final_excel_path))
            excel_ready_path = final_excel_path
            logger.info(f"Moved Excel to: {final_excel_path}")
        except Exception as e:
            logger.warning(f"Could not move Excel to OneDrive: {e}. Using local copy.")
        
        if pdf_success and temp_pdf_path.exists():
            try:
                shutil.move(str(temp_pdf_path), str(final_pdf_path))
                pdf_ready_path = final_pdf_path
                logger.info(f"Moved PDF to: {final_pdf_path}")
            except Exception as e:
                logger.warning(f"Could not move PDF to OneDrive: {e}. Using local copy.")
        else:
            logger.warning("PDF was not created successfully")
        
        # Move original CSV to converted folder
        converted_csv_path = CONVERTED_FOLDER / csv_file.name
        try:
            shutil.move(str(csv_file), str(converted_csv_path))
            logger.info(f"Moved original CSV to: {converted_csv_path}")
        except Exception as e:
            logger.warning(f"Could not archive original CSV: {e}")
        
        # Create email with attachments (use whatever paths we have)
        if excel_ready_path.exists() and (not pdf_success or pdf_ready_path.exists()):
            logger.info("")
            create_outlook_email(excel_ready_path, pdf_ready_path if pdf_ready_path.exists() else None, month_year)
        else:
            logger.warning("Skipping email creation - essential files missing")
        
        logger.info(f"✓ Successfully processed: {csv_file.name}")
        return True
        
    except Exception as e:
        logger.error(f"✗ Error processing {csv_file.name}: {e}", exc_info=True)
        return False


def main():
    """
    Main function to process all pending timesheet files.
    """
    logger.info("Starting timesheet processing...")
    logger.info(f"Import folder: {IMPORT_FOLDER}")
    logger.info(f"OneDrive folder: {ONEDRIVE_FOLDER}")
    
    # Find all CSV files matching the pattern
    csv_files = list(IMPORT_FOLDER.glob("*-LoggedTime-*.csv"))
    
    if not csv_files:
        logger.info("No files found to process.")
        logger.info(f"Place CSV files matching '*-LoggedTime-*.csv' in: {IMPORT_FOLDER}")
        return
    
    logger.info(f"Found {len(csv_files)} file(s) to process")
    
    # Process each file
    success_count = 0
    for csv_file in csv_files:
        if process_timesheet_file(csv_file):
            success_count += 1
    
    # Summary
    logger.info(f"\n{'='*60}")
    logger.info(f"Processing complete!")
    logger.info(f"Successfully processed: {success_count}/{len(csv_files)} files")
    logger.info(f"Log file: {log_file}")
    logger.info(f"{'='*60}")


if __name__ == "__main__":
    main()
