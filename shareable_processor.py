#!/usr/bin/env python3
"""
Shareable Timesheet Automation Script

Processes CSV files from Float/People and converts them to Excel and PDF formats.
Automatically organizes files into appropriate folders, configurable per user.
Built to be OS-agnostic and AI-friendly.

Usage:
    python shareable_processor.py --help
"""

import os
import sys
import re
import json
import logging
import argparse
import platform
import subprocess
import shutil
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Configure Base Directories
SCRIPT_DIR = Path(__file__).parent.absolute()
LOG_FOLDER = SCRIPT_DIR / "logs"
CONVERTED_FOLDER = SCRIPT_DIR / "converted"
CONFIG_FILE = SCRIPT_DIR / "config.json"

LOG_FOLDER.mkdir(exist_ok=True)
CONVERTED_FOLDER.mkdir(exist_ok=True)

# Setup Logging (JSON output friendly for AI if needed, but standard text is okay)
log_file = LOG_FOLDER / f"shareable_processing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

DEFAULT_CONFIG = {
    "employee_name": "Your Name",
    "client_name": "Client",
    "target_folder": str(Path.home() / "Documents" / "Urenstaat"),
    "email_enabled": True,
    "email_recipient": "uren@polteq.com",
    "email_cc": "",
    "email_subject": "Urenstaat {employee_name} - {client_name} {month_year}",
    "email_body": "Zie bijlage voor de urenstaat van {month_year}.\n\n"
}

def load_config(interactive=False):
    """Loads configuration, optionally prompting interactively if missing."""
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, 'r') as f:
                config = json.load(f)
                return config
        except Exception as e:
            logger.error(f"Failed to read config file: {e}")
    
    if interactive:
        print("\n--- First Time Setup ---")
        config = DEFAULT_CONFIG.copy()
        
        name = input(f"Employee Name [{config['employee_name']}]: ").strip()
        if name: config['employee_name'] = name
        
        client = input(f"Client Name [{config['client_name']}]: ").strip()
        if client: config['client_name'] = client
        
        folder = input(f"Target Output Folder (e.g. OneDrive) [{config['target_folder']}]: ").strip()
        if folder: config['target_folder'] = folder
        
        save_config(config)
        print("Configuration saved!\n")
        return config
    else:
        logger.warning("Config file missing. Generating default config.json.")
        save_config(DEFAULT_CONFIG)
        return DEFAULT_CONFIG

def save_config(config):
    with open(CONFIG_FILE, 'w') as f:
        json.dump(config, f, indent=4)
    logger.info(f"Saved configuration to {CONFIG_FILE}")

def extract_date_from_filename(filename):
    """Extract end date from filename pattern: *-LoggedTime-YYYYMMDD-YYYYMMDD.csv"""
    pattern = r'(.*?)-LoggedTime-(\d{8})-(\d{8}).*?\.csv'
    match = re.search(pattern, filename)
    
    if not match:
        raise ValueError(f"Filename '{filename}' does not match expected pattern (*-LoggedTime-YYYYMMDD-YYYYMMDD.csv)")
    
    end_date_str = match.group(3)
    end_date = datetime.strptime(end_date_str, '%Y%m%d')
    return end_date

def read_csv_flexibly(file_path):
    """Read CSV file, trying multiple delimiters."""
    for delimiter in [',', '\t', ';']:
        try:
            df = pd.read_csv(file_path, delimiter=delimiter)
            if len(df.columns) > 1:
                return df
        except Exception:
            pass
    try:
        df = pd.read_csv(file_path, sep=None, engine='python')
        return df
    except Exception as e:
        raise ValueError(f"Failed to read CSV with any delimiter: {e}")

def convert_to_excel(df, output_path):
    """Convert DataFrame to Excel file with basic formatting."""
    logger.info(f"Converting to Excel: {output_path}")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Timesheet')
        worksheet = writer.sheets['Timesheet']
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

def convert_to_pdf_pure_python(df, output_path, title):
    """Generate high-quality PDF using ReportLab without external dependencies."""
    logger.info(f"Converting to PDF (Pure Python): {output_path}")
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        
        doc = SimpleDocTemplate(str(output_path), pagesize=landscape(A4),
                                rightMargin=30, leftMargin=30,
                                topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=15
        )
        elements.append(Paragraph(f"<b>{title}</b>", title_style))
        elements.append(Spacer(1, 10))
        
        # Format dates if they exist, to fit better
        df_display = df.copy()
        # Convert all to string to avoid issues
        df_display = df_display.astype(str)
        
        # Replace NaN strings
        df_display = df_display.replace('nan', '')
        
        data = [df_display.columns.tolist()] + df_display.values.tolist()
        
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
        ]))
        
        elements.append(table)
        doc.build(elements)
        logger.info(f"PDF created successfully: {output_path}")
        return True
    except ImportError:
        logger.error("ReportLab not installed. Run 'pip install reportlab'.")
        return False
    except Exception as e:
        logger.error(f"Error creating PDF: {e}")
        return False

def convert_to_pdf_numbers(csv_path, pdf_path):
    """Convert CSV to PDF using Numbers.app (macOS native) for best formatting."""
    logger.info(f"Converting to PDF using Numbers.app: {pdf_path}")
    try:
        # Use absolute paths for AppleScript
        abs_csv = os.path.abspath(csv_path)
        abs_pdf = os.path.abspath(pdf_path)
        
        applescript = f'''
        tell application "Numbers"
            set theDoc to open POSIX file "{abs_csv}"
            delay 1
            export theDoc to POSIX file "{abs_pdf}" as PDF
            close theDoc without saving
        end tell
        '''
        
        result = subprocess.run(['osascript', '-e', applescript], capture_output=True, text=True, timeout=30)
        
        if result.returncode == 0 and os.path.exists(abs_pdf):
            logger.info(f"✓ PDF created successfully using Numbers: {pdf_path}")
            return True
        else:
            logger.warning(f"Numbers export failed: {result.stderr}")
            return False
    except Exception as e:
        logger.warning(f"Error using Numbers for PDF conversion: {e}")
        return False

def convert_to_pdf_excel_win32(excel_path, pdf_path):
    """Convert Excel to PDF using Microsoft Excel (Windows native) for best formatting."""
    logger.info(f"Converting to PDF using Excel (win32com): {pdf_path}")
    try:
        import win32com.client
        from pythoncom import com_error
        
        excel_abs = os.path.abspath(excel_path)
        pdf_abs = os.path.abspath(pdf_path)
        
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        try:
            wb = excel.Workbooks.Open(excel_abs)
            # 0 = xlTypePDF
            wb.ExportAsFixedFormat(0, pdf_abs)
            wb.Close(False)
            logger.info(f"✓ PDF created successfully using Excel: {pdf_path}")
            return True
        except com_error as e:
            logger.warning(f"Excel export failed: {e}")
            return False
        finally:
            excel.Quit()
    except ImportError:
        logger.warning("pywin32 not installed. Cannot use Excel for PDF conversion.")
        return False
    except Exception as e:
        logger.warning(f"Error using Excel for PDF conversion: {e}")
        return False

def create_outlook_email(config, excel_path, pdf_path, month_year):
    """Create a new email in Outlook, cross-platform."""
    if not config.get("email_enabled", True):
        logger.info("Email automation is disabled in config.")
        return False
    
    subject = config["email_subject"].format(
        employee_name=config["employee_name"],
        client_name=config["client_name"],
        month_year=month_year
    )
    body = config["email_body"].format(
        employee_name=config["employee_name"],
        client_name=config["client_name"],
        month_year=month_year
    )
    recipient = config.get("email_recipient", "")
    cc = config.get("email_cc", "")
    
    os_name = platform.system()
    
    if os_name == "Darwin":
        logger.info("Drafting email using macOS AppleScript...")
        try:
            applescript = f'''
            tell application "Microsoft Outlook"
                set newMessage to make new outgoing message with properties {{subject:"{subject}", content:"{body}"}}
                {f'make new recipient at newMessage with properties {{email address:{{address:"{recipient}"}}}}' if recipient else ''}
                {f'make new cc recipient at newMessage with properties {{email address:{{address:"{cc}"}}}}' if cc else ''}
                make new attachment at newMessage with properties {{file:POSIX file "{excel_path}"}}
                make new attachment at newMessage with properties {{file:POSIX file "{pdf_path}"}}
                open newMessage
                activate
            end tell
            '''
            subprocess.run(['osascript', '-e', applescript], capture_output=True, text=True, check=True)
            logger.info("✓ Outlook email draft created (macOS).")
            return True
        except Exception as e:
            logger.error(f"Failed macOS email creation: {e}")
            return False
            
    elif os_name == "Windows":
        logger.info("Drafting email using Windows win32com...")
        try:
            import win32com.client
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            mail.To = recipient
            mail.CC = cc
            mail.Subject = subject
            mail.Body = body
            mail.Attachments.Add(str(excel_path))
            mail.Attachments.Add(str(pdf_path))
            mail.Display(True)
            logger.info("✓ Outlook email draft created (Windows).")
            return True
        except ImportError:
            logger.error("pywin32 not installed. Run 'pip install pywin32'.")
            return False
        except Exception as e:
            logger.error(f"Failed Windows email creation: {e}")
            return False
    else:
        logger.warning(f"OS {os_name} not supported for automated email drafting.")
        return False

def process_file(csv_file_path, config):
    """Core logic to process a single timesheet file."""
    csv_path = Path(csv_file_path)
    if not csv_path.exists():
        logger.error(f"File not found: {csv_path}")
        return False

    logger.info(f"Processing: {csv_path.name}")
    try:
        end_date = extract_date_from_filename(csv_path.name)
        month_year = end_date.strftime('%m-%Y')
        
        df = read_csv_flexibly(csv_path)
        logger.info(f"Loaded {len(df)} rows.")
        
        base_filename = f"Urenstaat {config['client_name']}-{config['employee_name']} {month_year}"
        excel_filename = f"{base_filename}.xlsx"
        pdf_filename = f"{base_filename}.pdf"
        
        target_folder = Path(config["target_folder"])
        target_folder.mkdir(parents=True, exist_ok=True)
        
        excel_path = target_folder / excel_filename
        pdf_path = target_folder / pdf_filename
        
        convert_to_excel(df, excel_path)
        
        # Try native conversion on macOS if possible
        pdf_success = False
        os_name = platform.system()
        if os_name == "Darwin":
            pdf_success = convert_to_pdf_numbers(csv_path, pdf_path)
        elif os_name == "Windows":
            pdf_success = convert_to_pdf_excel_win32(excel_path, pdf_path)
        
        # Fallback for any OS
        if not pdf_success:
            pdf_success = convert_to_pdf_pure_python(df, pdf_path, title=base_filename)
        
        # Archive original CSV
        archived_csv_path = CONVERTED_FOLDER / csv_path.name
        # Using copy + unlink to avoid cross-device link issues if folders are on different partitions
        shutil.copy2(str(csv_path), str(archived_csv_path))
        os.remove(str(csv_path))
        logger.info(f"Archived original CSV to {archived_csv_path}")
        
        # Draft email
        create_outlook_email(config, excel_path, pdf_path, month_year)
        
        logger.info(f"✓ Successfully processed {csv_path.name}")
        return True

    except Exception as e:
        logger.error(f"Error processing {csv_path.name}: {e}", exc_info=True)
        return False

def main():
    parser = argparse.ArgumentParser(description="Shareable Timesheet Processor")
    parser.add_argument("files", nargs="*", help="CSV files to process")
    parser.add_argument("--setup", action="store_true", help="Run interactive setup to configure the tool")
    
    args = parser.parse_args()
    
    if args.setup:
        load_config(interactive=True)
        return

    config = load_config(interactive=False)
    
    # Process files provided via CLI arguments
    if args.files:
        for file in args.files:
            process_file(file, config)
        return
        
    # If no files provided, check 'Import Urenstaat' folder as fallback
    legacy_import_folder = SCRIPT_DIR / "Import Urenstaat"
    if legacy_import_folder.exists():
        csv_files = list(legacy_import_folder.glob("*-LoggedTime-*.csv"))
        if csv_files:
            logger.info(f"Found {len(csv_files)} file(s) in {legacy_import_folder}")
            for file in csv_files:
                process_file(file, config)
            return

    logger.info("No files provided or found. Usage: drag & drop a file, or run with file arguments.")

if __name__ == "__main__":
    main()
